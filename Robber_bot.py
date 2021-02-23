from twitchio.ext import commands
import openpyxl
from openpyxl import load_workbook
from decouple import config

temp_initial_channels = str(config('CHANNEL'))

class Bot(commands.Bot):
    def __init__(self):
        super().__init__(irc_token=config('TMI_TOKEN'), client_id=config('CLIENT_ID'), nick=config('BOT_NICK'), prefix=config('BOT_PREFIX'),
                         initial_channels=[temp_initial_channels])

    async def event_ready(self):
        ws = bot._ws
        await ws.send_privmsg("whitebot201", f"whitebot201 이 준비 완료 되었습니다!")

    async def event_message(self, message):
        print(message.content)
        await self.handle_commands(message)
    
    @commands.command(name='명령어')
    async def 명령어(self, ctx):
        await ctx.send("?보상추가, ?보상목록, ?보상삭제, ?유저삭제")

    @commands.command(name='test')
    async def test(self, ctx):
        await ctx.send(f'test pass!!!')

    @commands.command(name='보상추가')
    async def 보상추가(self, ctx, user_name = None, user_reward = None):
        print("닉네임과 추가할 보상을 확인하는 중 [닉네임 : "+str(user_name)+" / 보상 : "+ str(user_reward)+"]")
        if user_name == None or user_reward == None:
            if user_name == None and user_reward == None:
                await ctx.send("?보상추가 [닉네임] [추가할보상] 의 형식으로 입력해주십시오.")
                await ctx.send("[닉네임] 과 [추가할보상] 을 입력하지 않았습니다.")
            elif user_name != None and user_reward == None:
                await ctx.send("?보상추가 [닉네임] [추가할보상] 의 형식으로 입력해주십시오.")
                await ctx.send("[추가할보상] 을 입력하지 않았습니다.")
            elif user_name == None and user_reward != None:
                await ctx.send("?보상추가 [닉네임] [추가할보상] 의 형식으로 입력해주십시오.")
                await ctx.send("[닉네임] 을 입력하지 않았습니다.")
            else:
                await ctx.send("?보상추가 [닉네임] [추가할보상] 의 형식으로 입력해주십시오.")    
        elif user_name != None and user_reward != None:
            print(1)
            file = openpyxl.load_workbook("Robber_data.xlsx")
            sheet = file.active
            i = 1
            while True:
                if sheet["A" + str(i)].value == user_name:
                    j = 65
                    while True:
                        print(sheet[chr(j) + str(i)].value)
                        if sheet[chr(j) + str(i)].value == None:
                            sheet[chr(j) + str(i)].value = user_reward
                            break
                        j=j+1
                    break
                elif sheet["A" + str(i)].value == None:
                    sheet["A" + str(i)].value = user_name
                    sheet["B" + str(i)].value = user_reward
                    break
                i=i+1
            file.save("Robber_data.xlsx")
            await ctx.send(str(user_name)+"의 보상["+str(user_reward)+"]을 추가 하였습니다.")

    @commands.command(name='보상목록')
    async def 보상목록(self, ctx, user_name=None):
        print("입력 받은 유저의 닉네임은 "+str(user_name)+" 입니다.")
        file = openpyxl.load_workbook("Robber_data.xlsx")
        sheet = file.active
        if user_name == None:
            print("유저 닉네임을 입력받지 못해 전체 유저 리스트를 출력 시작 중...")#
            user_reward_list=[]
            i=1
            while True:
                if sheet["A" + str(i)].value == None:
                    print("전체 유저 리스트의 마지막을 찾는 중...")#
                    break
                elif sheet["A" + str(i)].value != None and sheet["A" + str(i)].value !=None and sheet["B" + str(i)].value !=None:
                    print("전체 유저 리스트 하나씩 출력 중...")#
                    user_reward_list.append(str(sheet["A" + str(i)].value))
                i=i+1
            msg = ""
            print("출력할 전체 유저 리스트의 형식을 지정 중...")#
            for i in range(len(user_reward_list)):
                if i!=0:
                    msg +=", "
                msg += str(user_reward_list[i])
            
            await ctx.send("보상이 있는 유저의 목록은 다음과 같습니다.\n["+str(msg)+"]. ?보상목록 [닉네임] 의 형식으로 입력해주세요. ")
            await ctx.send("(보상이 없는 유저도 포함해서 출력하고 싶으시다면 ?보상목록 전체 혹은 ?보상목록 모두를 해주세요)")

        elif user_name=="전체" or user_name=="모두":
            print("유저 닉네임을 입력받지 못해 전체 유저 리스트를 출력 시작 중...")#
            user_reward_list=[]
            i=1
            while True:
                if sheet["A" + str(i)].value == None:
                    print("전체 유저 리스트의 마지막을 찾는 중...")#
                    break
                elif sheet["A" + str(i)].value != None and sheet["A" + str(i)].value !=None:
                    print("전체 유저 리스트 하나씩 출력 중...")#
                    user_reward_list.append(str(sheet["A" + str(i)].value))
                i=i+1
            msg = ""
            print("출력할 전체 유저 리스트의 형식을 지정 중...")#
            for i in range(len(user_reward_list)):
                if i!=0:
                    msg +=", "
                msg += user_reward_list[i]
            
            await ctx.send("보상 목록에 있는 유저는 다음과 같습니다.\n["+str(msg)+"].\n?보상목록 [닉네임] 의 형식으로 입력해주세요.")

        elif user_name!=None and user_name!="모두"and user_name!="전체":
            print("유저 닉네임을 입력 받아 해당 유저를 보상 리스트에서 찾는 중...")#
            file = openpyxl.load_workbook("Robber_data.xlsx")
            sheet = file.active
            msg = ""
            i=1
            while True:
                print(sheet["A" + str(i)].value)
                print(111)
                if sheet["A" + str(i)].value == user_name:
                    msg += user_name+" 님의 보상은 "
                    j=66
                    print("입력 받은 유저 닉네임을 보상 리스트에서 찾음...")#
                    while True:
                        if j==66 and sheet[str(chr(j)) + str(i)].value == None:
                            msg += "[] (보상이 없는 것 같습니다. 관리자의 확인을 요청하십시오)   "#
                            break
                        elif sheet[str(chr(j)) + str(i)].value == None:
                            print("보상 리스트에서 마지막 항목을 찾음...")#
                            break
                        else:
                            print("보상 리스트에서 항목을 찾음...")#
                            msg += "["+sheet[str(chr(j)) + str(i)].value+"]"
                            msg += ", "
                        j+=1
                    await ctx.send(msg[:-2])
                    break
                elif sheet["A" + str(i)].value == None:
                    print("보상 목록에서 입력 받은 유저 닉네임을 찾지 못하는 중...")
                    await ctx.send("보상목록에서 "+user_name+" 님을 찾을 수 없었습니다.")
                    break
                i+=1
            file.save("Robber_data.xlsx")

    @commands.command(name='보상삭제')
    async def 보상삭제(self, ctx,user_name = None, delete_reward = None):
        if user_name == None or delete_reward == None:
            await ctx.send("?보상추가 [닉네임] [삭제할보상] 의 형식으로 입력해주십시오.")
            if user_name == None and delete_reward == None:
                await ctx.send("[닉네임] 과 [추가할보상] 을 입력하지 않았습니다.")
            elif user_name != None and delete_reward == None:
                await ctx.send("[삭제할보상] 을 입력하지 않았습니다.")
            elif user_name == None and delete_reward != None:
                await ctx.send("[닉네임] 을 입력하지 않았습니다.")
        else:
            print("보상 삭제 시작중...")
            file = openpyxl.load_workbook("Robber_data.xlsx")
            sheet = file.active
            user_num = 0
            max_num = 0
            i=1
            while True:
                print("보상을 삭제할 유저를 찾는중...")
                if sheet["A" + str(i)].value == user_name:
                    user_num = i
                    j = 66
                    while True:
                        print("보상을 삭제할 유저의 총 보상 개수를 구하는중...")
                        if sheet[chr(j) + str(user_num)].value == None:
                            max_num = j-1
                            break
                        j+=1
                    break
                if sheet["A" + str(i)].value == None:
                    await ctx.send(user_name+" 님이 보상 목록에 없습니다. !보상목록으로 닉네임을 다시 한번 확인해주십시오")
                    break
                i+=1
            if user_num !=0:
                i = 66
                while True:
                    print("삭제할 보상을 찾는 중...")
                    if sheet[str(chr(i)) + str(user_num)].value == None or  i>90:
                        print("삭제할 보상을 찾지 못하는 중...")
                        await ctx.send(user_name+" 님의 보상 목록 중에서 " +delete_reward+ " 가 없습니다. !보상목록으로 보상을 다시 한번 확인해주십시오")
                        break
                    elif str(sheet[str(chr(i)) + str(user_num)].value) == str(delete_reward):
                        print("보상을 삭제 중...")
                        sheet[chr(i) + str(user_num)].value = ""
                        sheet[chr(i) + str(user_num)].value = sheet[chr(max_num) + str(user_num)].value
                        sheet[chr(max_num) + str(user_num)].value = ""
                        await ctx.send(user_name+" 님의 보상 목록 중에서 " +delete_reward+ " 을 삭제하였습니다.")
                        break
                    
                    i+=1
            file.save("Robber_data.xlsx")


    @commands.command(name='유저삭제')
    async def 유저삭제(self, ctx, user_name=None):
        file = openpyxl.load_workbook("Robber_data.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value==None:
                await ctx.send(user_name+" 님을 보상목록에서 찾지 못하였습니다. ?보상목록 명령어로 다시 확인해주십시오.")
                break
            elif str(sheet["A" + str(i)].value)==user_name and sheet["B" + str(i)].value==None:
                print(i)
                sheet.delete_rows(i)
                await ctx.send(user_name+" 님을 보상목록에서 삭제하였습니다.")
                break
            elif str(sheet["A" + str(i)].value)==user_name and sheet["B" + str(i)].value!=None:
                await ctx.send(user_name+" 님의 보상이 남아있어서 보상목록에서 "+user_name+" 님을 삭제할 수 없습니다.")
                break
            i+=1

        file.save("Robber_data.xlsx")    

bot = Bot()
bot.run()